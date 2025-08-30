#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Controlador de Stock (Tkinter + SQLite)
Autor: Tú + ChatGPT

Descripción
-----------
App de escritorio en Python (Tkinter) con base SQLite local.
Incluye los componentes del diseño solicitado:

- Buscar [Entry]
- Botones: Nuevo Producto, Editar Producto, Eliminar Producto,
          Agregar Stock, Restar Stock, Movimientos, Refrescar, Exportar a Excel
- Campo "Escanear / Código" [Entry] que dispara salida rápida de stock (OUT con Enter)
- Tabla (Treeview): Código, Nombre, Precio Unitario, Stock

Movimientos registrados con tipo IN/OUT, cantidad y precio unitario.
Soporta ~2000 productos sin problema (índices en código y nombre).

Requisitos
----------
- Python 3.9+ (Tkinter viene incluido con la instalación estándar)
- Para exportar a **Excel (.xlsx)** se recomienda tener instalado **openpyxl**.
  Si no está instalado, la app ofrece exportar a CSV como alternativa.

Ejecución
---------
python stock_app.py

Compilar a .exe (opcional, Windows)
-----------------------------------
pip install pyinstaller
pyinstaller --noconfirm --onefile --name ControladorStock stock_app.py

Notas
-----
- La base se guarda por defecto en:
    Windows: %LOCALAPPDATA%/OnceStock/inventario.db
    Otros SO: ./inventario.db (carpeta actual)
- Podés forzar la ruta usando la variable de entorno STOCK_DB.
"""

import os
import sqlite3
import sys
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from datetime import datetime
import csv
import re

try:
    import openpyxl
    from openpyxl.styles import Font
    import openpyxl.utils
except Exception:
    openpyxl = None

APP_NAME = "Controlador de Stock"
VERSION = "1.2.0"

def _db_default_path() -> str:
    env = os.environ.get("STOCK_DB")
    if env:
        return env
    if os.name == "nt":
        root = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
        base_dir = os.path.join(root, "OnceStock")
        os.makedirs(base_dir, exist_ok=True)
        return os.path.join(base_dir, "inventario.db")
    else:
        return os.path.abspath("./inventario.db")

DB_PATH = _db_default_path()

SCHEMA = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS productos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    codigo TEXT NOT NULL UNIQUE,
    nombre TEXT NOT NULL,
    precio REAL NOT NULL DEFAULT 0,
    stock INTEGER NOT NULL DEFAULT 0,
    barcode TEXT UNIQUE,
    creado_en TEXT NOT NULL,
    actualizado_en TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS movimientos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    producto_id INTEGER NOT NULL,
    tipo TEXT NOT NULL CHECK(tipo IN ('IN','OUT')),
    cantidad INTEGER NOT NULL CHECK(cantidad > 0),
    precio_unitario REAL NOT NULL CHECK(precio_unitario >= 0),
    nota TEXT,
    creado_en TEXT NOT NULL,
    FOREIGN KEY(producto_id) REFERENCES productos(id) ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_productos_nombre ON productos(nombre);
CREATE INDEX IF NOT EXISTS idx_productos_codigo ON productos(codigo);
"""

def now_str() -> str:
    return datetime.now().isoformat(timespec="seconds")

def parse_number(s: str, default=0.0) -> float:
    """Convierte '1.600,50' o '1,600.50' o '$ 1.600' a float seguro."""
    if s is None:
        return float(default)
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    if not s:
        return float(default)
    # remove currency and spaces
    s = s.replace("$","").replace("ARS","").replace("usd","").replace("USD","").strip()
    # keep digits and separators
    s = re.sub(r"[^0-9\.,-]", "", s)
    if s.count(",") and s.count("."):
        # decide last separator as decimal
        last_dot = s.rfind(".")
        last_com = s.rfind(",")
        if last_dot > last_com:
            # dot is decimal, remove commas
            s = s.replace(",", "")
        else:
            # comma is decimal: remove dots and replace comma with dot
            s = s.replace(".", "").replace(",", ".")
    else:
        # only one type or none
        if s.count(",") and not s.count("."):
            s = s.replace(",", ".")
        # if only dots, leave as is
    try:
        return float(s)
    except Exception:
        return float(default)

class DB:
    def __init__(self, path: str):
        self.path = path
        self._ensure()

    def connect(self):
        conn = sqlite3.connect(self.path)
        conn.row_factory = sqlite3.Row
        return conn

    def _ensure(self):
        with self.connect() as conn:
            conn.executescript(SCHEMA)
            # seed si está vacío
            cur = conn.execute("SELECT COUNT(*) AS c FROM productos")
            if cur.fetchone()["c"] == 0:
                seed = [
                    ("94319699", "billy", 1600.00, 40, None),
                    ("56070724", "evan", 1600.00, 30, None),
                    ("94466555", "shay",    0.00, 20, None),
                ]
                for codigo, nombre, precio, stock, barcode in seed:
                    conn.execute(
                        "INSERT INTO productos (codigo, nombre, precio, stock, barcode, creado_en, actualizado_en) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (codigo, nombre, precio, stock, barcode, now_str(), now_str())
                    )

    # Productos
    def listar_productos(self, filtro:str=""):
        q = """
        SELECT id, codigo, nombre, precio, stock
        FROM productos
        WHERE (? = '' OR codigo LIKE ? OR nombre LIKE ?)
        ORDER BY nombre COLLATE NOCASE
        """
        like = f"%{filtro.strip()}%"
        with self.connect() as conn:
            return conn.execute(q, (filtro.strip(), like, like)).fetchall()

    def crear_producto(self, codigo, nombre, precio, stock):
        with self.connect() as conn:
            conn.execute(
                "INSERT INTO productos (codigo, nombre, precio, stock, barcode, creado_en, actualizado_en) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (codigo, nombre, float(precio), int(stock), None, now_str(), now_str())
            )

    def obtener_producto(self, pid):
        with self.connect() as conn:
            return conn.execute("SELECT * FROM productos WHERE id=?", (pid,)).fetchone()

    def obtener_producto_por_codigo(self, codigo):
        with self.connect() as conn:
            return conn.execute("SELECT * FROM productos WHERE codigo=?", (codigo,)).fetchone()

    def actualizar_producto(self, pid, **fields):
        if not fields:
            return
        fields["actualizado_en"] = now_str()
        cols = ", ".join(f"{k}=?" for k in fields.keys())
        vals = list(fields.values()) + [pid]
        with self.connect() as conn:
            conn.execute(f"UPDATE productos SET {cols} WHERE id=?", vals)

    def eliminar_producto(self, pid):
        with self.connect() as conn:
            conn.execute("DELETE FROM productos WHERE id=?", (pid,))

    # Movimientos
    def crear_movimiento(self, producto_id, tipo, cantidad, precio_unitario, nota=None):
        cantidad = int(cantidad)
        precio_unitario = float(precio_unitario)
        if cantidad <= 0 or precio_unitario < 0:
            raise ValueError("Cantidad o precio inválidos")
        with self.connect() as conn:
            conn.execute("""
                INSERT INTO movimientos (producto_id, tipo, cantidad, precio_unitario, nota, creado_en)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (producto_id, tipo, cantidad, precio_unitario, nota, now_str()))
            mult = 1 if tipo == "IN" else -1
            conn.execute("UPDATE productos SET stock = stock + ?, actualizado_en=? WHERE id=?",
                         (mult * cantidad, now_str(), producto_id))

    def listar_movimientos(self, producto_id=None, limite=500):
        if producto_id:
            q = """
            SELECT m.id, p.codigo, p.nombre, m.tipo, m.cantidad, m.precio_unitario, m.nota, m.creado_en
            FROM movimientos m
            JOIN productos p ON p.id = m.producto_id
            WHERE producto_id=?
            ORDER BY m.id DESC
            LIMIT ?
            """
            args = (producto_id, limite)
        else:
            q = """
            SELECT m.id, p.codigo, p.nombre, m.tipo, m.cantidad, m.precio_unitario, m.nota, m.creado_en
            FROM movimientos m
            JOIN productos p ON p.id = m.producto_id
            ORDER BY m.id DESC
            LIMIT ?
            """
            args = (limite,)
        with self.connect() as conn:
            return conn.execute(q, args).fetchall()


class ProductoForm(tk.Toplevel):
    def __init__(self, master, db:DB, producto=None):
        super().__init__(master)
        self.db = db
        self.producto = producto
        self.title("Nuevo Producto" if producto is None else "Editar Producto")
        self.resizable(False, False)
        self.grab_set()

        # Campos
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        ttk.Label(frm, text="Código*").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        self.e_codigo = ttk.Entry(frm, width=30)
        self.e_codigo.grid(row=0, column=1, sticky="w", padx=6, pady=6)

        ttk.Label(frm, text="Nombre*").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        self.e_nombre = ttk.Entry(frm, width=40)
        self.e_nombre.grid(row=1, column=1, sticky="w", padx=6, pady=6)

        ttk.Label(frm, text="Precio Unitario").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        self.e_precio = ttk.Entry(frm, width=20)
        self.e_precio.insert(0, "0")
        self.e_precio.grid(row=2, column=1, sticky="w", padx=6, pady=6)

        ttk.Label(frm, text="Stock Inicial").grid(row=3, column=0, sticky="e", padx=6, pady=6)
        self.e_stock = ttk.Entry(frm, width=20)
        self.e_stock.insert(0, "0")
        self.e_stock.grid(row=3, column=1, sticky="w", padx=6, pady=6)

        btns = ttk.Frame(frm)
        btns.grid(row=4, column=0, columnspan=2, pady=(12,0))
        ttk.Button(btns, text="Guardar", command=self._guardar).pack(side="left", padx=6)
        ttk.Button(btns, text="Cancelar", command=self.destroy).pack(side="left", padx=6)

        if producto:
            self.e_codigo.insert(0, producto["codigo"])
            self.e_nombre.insert(0, producto["nombre"])
            self.e_precio.delete(0, "end"); self.e_precio.insert(0, f'{producto["precio"]:.2f}')
            self.e_stock.delete(0, "end"); self.e_stock.insert(0, str(producto["stock"]))

    def _guardar(self):
        codigo = self.e_codigo.get().strip()
        nombre = self.e_nombre.get().strip()
        try:
            precio = parse_number(self.e_precio.get().strip() or "0")
        except Exception:
            messagebox.showerror("Error", "Precio inválido")
            return
        try:
            stock = int(float(self.e_stock.get().strip() or 0))
        except ValueError:
            messagebox.showerror("Error", "Stock inválido")
            return

        if not codigo or not nombre:
            messagebox.showerror("Error", "Código y Nombre son obligatorios")
            return

        try:
            if self.producto is None:
                self.db.crear_producto(codigo, nombre, precio, stock)
            else:
                self.db.actualizar_producto(self.producto["id"], codigo=codigo, nombre=nombre, precio=precio, stock=stock)
        except sqlite3.IntegrityError as e:
            messagebox.showerror("Error", f"Código duplicado.\n\n{e}")
            return
        self.destroy()


class MovimientosView(tk.Toplevel):
    def __init__(self, master, db:DB, producto_id=None):
        super().__init__(master)
        self.db = db
        self.title("Movimientos")
        self.geometry("800x400")
        self.grab_set()

        cols = ("id","fecha","codigo","nombre","tipo","cantidad","precio","nota")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        self.tree.heading("id", text="ID")
        self.tree.heading("fecha", text="Fecha")
        self.tree.heading("codigo", text="Código")
        self.tree.heading("nombre", text="Nombre")
        self.tree.heading("tipo", text="Tipo")
        self.tree.heading("cantidad", text="Cant.")
        self.tree.heading("precio", text="P.Unit.")
        self.tree.heading("nota", text="Nota")
        self.tree.column("id", width=60, anchor="e")
        self.tree.column("fecha", width=140)
        self.tree.column("codigo", width=110)
        self.tree.column("nombre", width=180)
        self.tree.column("tipo", width=60, anchor="center")
        self.tree.column("cantidad", width=70, anchor="e")
        self.tree.column("precio", width=90, anchor="e")
        self.tree.column("nota", width=200)
        self.tree.pack(fill="both", expand=True, padx=8, pady=8)

        self._load(producto_id)

    def _load(self, producto_id=None):
        for i in self.tree.get_children():
            self.tree.delete(i)
        rows = self.db.listar_movimientos(producto_id=producto_id, limite=1000)
        for r in rows:
            self.tree.insert("", "end", values=(
                r["id"],
                r["creado_en"],
                r["codigo"],
                r["nombre"],
                r["tipo"],
                r["cantidad"],
                f"${r['precio_unitario']:,.2f}",
                r["nota"] or ""
            ))


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME} — v{VERSION}")
        self.geometry("1250x600")  # Tamaño intermedio para fuente de 12 puntos
        self.db = DB(DB_PATH)
        self._setup_style()
        self._build_ui()
        self._load_table()

    def _setup_style(self):
        style = ttk.Style(self)
        try:
            # Scaling moderado para fuente de 12 puntos
            self.tk.call("tk", "scaling", 1.3)
            
            # Configurar fuentes personalizadas
            self.default_font = ("Segoe UI", 12)
            self.button_font = ("Segoe UI", 12)
            self.entry_font = ("Segoe UI", 12)
            
            # Aplicar fuentes a los estilos de ttk
            style.configure("TLabel", font=self.default_font)
            style.configure("TButton", font=self.button_font)
            style.configure("TEntry", font=self.entry_font)
            style.configure("Treeview", font=("Segoe UI", 11))
            style.configure("Treeview.Heading", font=("Segoe UI", 12, "bold"))
            
        except Exception:
            pass

    # ----- UI -----
    def _build_ui(self):
        top = ttk.Frame(self, padding=(8,6,8,0))
        top.pack(fill="x")

        # Buscar
        ttk.Label(top, text="Buscar:").pack(side="left", padx=(0,6))
        self.e_buscar = ttk.Entry(top, width=30)
        self.e_buscar.pack(side="left")
        self.e_buscar.bind("<KeyRelease>", lambda e: self._load_table())
        self.e_buscar.delete(0, "end")  # Asegurar que esté vacío

        # Botonera principal
        ttk.Button(top, text="Nuevo Producto", command=self._nuevo_producto).pack(side="left", padx=6)
        ttk.Button(top, text="Editar Producto", command=self._editar_producto).pack(side="left", padx=6)
        ttk.Button(top, text="Eliminar Producto", command=self._eliminar_producto).pack(side="left", padx=6)

        ttk.Separator(top, orient="vertical").pack(side="left", fill="y", padx=8, pady=2)

        ttk.Button(top, text="Agregar Stock", command=lambda: self._mov_stock("IN")).pack(side="left", padx=6)
        ttk.Button(top, text="Restar Stock", command=lambda: self._mov_stock("OUT")).pack(side="left", padx=6)
        ttk.Button(top, text="Movimientos", command=self._abrir_movimientos).pack(side="left", padx=6)

        ttk.Separator(top, orient="vertical").pack(side="left", fill="y", padx=8, pady=2)

        ttk.Label(top, text="Escanear / Código:").pack(side="left", padx=(0,6))
        self.e_scan = ttk.Entry(top, width=24)
        self.e_scan.pack(side="left")
        self.e_scan.bind("<Return>", self._scan_enter)
        self.e_scan.delete(0, "end")  # Asegurar que esté vacío

        ttk.Button(top, text="Refrescar", command=self._refrescar).pack(side="left", padx=6)
        ttk.Button(top, text="Exportar a Excel", command=self._exportar_excel).pack(side="left", padx=6)

        # Tabla
        cols = ("codigo","nombre","precio","stock")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        self.tree.heading("codigo", text="Código")
        self.tree.heading("nombre", text="Nombre")
        self.tree.heading("precio", text="Precio Unitario")
        self.tree.heading("stock", text="Stock")
        self.tree.column("codigo", width=160, anchor="w")
        self.tree.column("nombre", width=420, anchor="w")
        self.tree.column("precio", width=150, anchor="e")
        self.tree.column("stock", width=90, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=8, pady=8)

        # Barra de estado
        self.status = tk.StringVar(value=f"Base: {DB_PATH}")
        statusbar = ttk.Label(self, textvariable=self.status, anchor="w", padding=(8,4))
        statusbar.pack(fill="x", side="bottom")

        # Doble click edita
        self.tree.bind("<Double-1>", lambda e: self._editar_producto())

    def _get_selected_product_id(self):
        sel = self.tree.selection()
        if not sel:
            return None
        item = self.tree.item(sel[0])
        codigo = item["values"][0]
        prod = self.db.obtener_producto_por_codigo(str(codigo))
        return prod["id"] if prod else None

    def _load_table(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        filtro = self.e_buscar.get().strip()
        rows = self.db.listar_productos(filtro=filtro)
        total = 0
        for r in rows:
            total += 1
            self.tree.insert("", "end", values=(
                r["codigo"],
                r["nombre"],
                f"${r['precio']:,.2f}",
                r["stock"],
            ))
        self.status.set(f"{total} productos — Base: {DB_PATH}")

    def _refrescar(self):
        """Refrescar la aplicación: limpiar campos y recargar tabla"""
        # Limpiar campos de entrada
        self.e_buscar.delete(0, "end")
        self.e_scan.delete(0, "end")
        # Recargar la tabla
        self._load_table()

    # Acciones
    def _nuevo_producto(self):
        dlg = ProductoForm(self, self.db, producto=None)
        self.wait_window(dlg)
        self._load_table()

    def _editar_producto(self):
        pid = self._get_selected_product_id()
        if not pid:
            messagebox.showinfo("Editar", "Seleccioná un producto de la tabla.")
            return
        producto = self.db.obtener_producto(pid)
        dlg = ProductoForm(self, self.db, producto=producto)
        self.wait_window(dlg)
        self._load_table()

    def _eliminar_producto(self):
        pid = self._get_selected_product_id()
        if not pid:
            messagebox.showinfo("Eliminar", "Seleccioná un producto de la tabla.")
            return
        prod = self.db.obtener_producto(pid)
        if not prod:
            return
        if messagebox.askyesno("Confirmar", f"¿Eliminar '{prod['nombre']}' (código {prod['codigo']})?"):
            try:
                self.db.eliminar_producto(pid)
                self._load_table()
            except sqlite3.IntegrityError as e:
                messagebox.showerror("Error", f"No se pudo eliminar.\n{e}")

    def _mov_stock(self, tipo):
        pid = self._get_selected_product_id()
        if not pid:
            messagebox.showinfo("Movimiento", "Seleccioná un producto de la tabla.")
            return
        prod = self.db.obtener_producto(pid)
        if not prod:
            return
        titulo = "Agregar Stock" if tipo == "IN" else "Restar Stock"
        try:
            cant = simpledialog.askinteger(titulo, "Cantidad:", minvalue=1, parent=self)
            if cant is None:
                return
            # Usar el precio actual del producto y sin nota
            precio = float(prod["precio"])
            self.db.crear_movimiento(pid, tipo, cant, precio, None)
            self._load_table()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _abrir_movimientos(self):
        pid = self._get_selected_product_id()
        MovimientosView(self, self.db, producto_id=pid)

    def _scan_enter(self, event=None):
        codigo = self.e_scan.get().strip()
        if not codigo:
            return
        prod = self.db.obtener_producto_por_codigo(codigo)
        if not prod:
            messagebox.showinfo("Producto no encontrado", f"No existe el código {codigo}.")
            self.e_scan.delete(0,"end")
            return
        
        # Mostrar opciones para el producto encontrado
        try:
            self._mostrar_opciones_producto(prod)
        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            self.e_scan.delete(0, "end")
    
    def _mostrar_opciones_producto(self, prod):
        """Muestra opciones para un producto escaneado"""
        import tkinter as tk
        from tkinter import ttk
        
        # Crear ventana personalizada
        dialog = tk.Toplevel(self)
        dialog.title("Opciones de Producto")
        dialog.geometry("400x240")  # Tamaño intermedio para fuente de 12 puntos
        dialog.resizable(False, False)
        dialog.grab_set()
        dialog.transient(self)
        
        # Centrar la ventana
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Marco principal
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        # Información del producto
        info_text = f"Código: {prod['codigo']}\nNombre: {prod['nombre']}\nStock actual: {prod['stock']}"
        ttk.Label(main_frame, text=info_text, font=('Segoe UI', 12)).pack(pady=(0, 20))
        
        # Marco para botones
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        
        def agregar_stock():
            dialog.destroy()
            self._ejecutar_movimiento_escaneado(prod, "IN")
        
        def restar_stock():
            dialog.destroy()
            self._ejecutar_movimiento_escaneado(prod, "OUT")
        
        def cancelar():
            dialog.destroy()
        
        # Botones
        ttk.Button(btn_frame, text="Agregar Stock", command=agregar_stock, width=16).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Restar Stock", command=restar_stock, width=16).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=cancelar, width=16).pack(side="left", padx=5)
        
        # Foco en el primer botón
        btn_frame.winfo_children()[0].focus()
    
    def _ejecutar_movimiento_escaneado(self, prod, tipo):
        """Ejecuta movimiento de stock para producto escaneado"""
        titulo = "Agregar Stock" if tipo == "IN" else "Restar Stock"
        try:
            cant = simpledialog.askinteger(titulo, "Cantidad:", minvalue=1, parent=self)
            if cant is None:
                return
            precio = float(prod["precio"])
            nota = "Movimiento por escaneo"
            self.db.crear_movimiento(prod["id"], tipo, cant, precio, nota)
            self._load_table()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _exportar_excel(self):
        path = filedialog.asksaveasfilename(
            title="Exportar productos a Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Todos los archivos","*.*")],
            initialfile="productos.xlsx"
        )
        if not path:
            return

        rows = self.db.listar_productos(filtro=self.e_buscar.get().strip())

        if openpyxl is None:
            if messagebox.askyesno(
                "openpyxl no instalado",
                "Para Excel (.xlsx) se necesita openpyxl.\n\n¿Querés exportar en CSV como alternativa?"
            ):
                csv_path = os.path.splitext(path)[0] + ".csv"
                try:
                    with open(csv_path, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f)
                        w.writerow(["codigo","nombre","precio","stock"])
                        for r in rows:
                            w.writerow([r["codigo"], r["nombre"], f"{r['precio']:.2f}", r["stock"]])
                    messagebox.showinfo("Exportación completada", f"Archivo CSV guardado en:\n{csv_path}\n\nPara .xlsx instalá:\npip install openpyxl")
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo exportar CSV:\n{e}")
            return

        # Exportar a XLSX con openpyxl
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Productos"

            headers = ["Código", "Nombre", "Precio Unitario", "Stock"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            for r in rows:
                ws.append([r["codigo"], r["nombre"], float(r["precio"]), int(r["stock"])])

            # Formatos y anchos de columna
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = u'"$"#,##0.00'

            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    cell.number_format = u'#,##0'

            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

            wb.save(path)
            messagebox.showinfo("Exportación completada", f"Archivo Excel guardado en:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar a Excel:\n{e}")


def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
